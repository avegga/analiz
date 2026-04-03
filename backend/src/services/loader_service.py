from datetime import datetime
from pathlib import Path

import openpyxl
import xlrd
from bs4 import BeautifulSoup

from src.models.template_registry import Template


def _decode_html_bytes(raw_bytes: bytes) -> str:
    for encoding in ("utf-8-sig", "cp1251", "utf-16", "latin-1"):
        try:
            return raw_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw_bytes.decode("utf-8", errors="replace")


def _looks_like_html(raw_bytes: bytes) -> bool:
    prefix = raw_bytes[:512].lstrip().removeprefix(b"\xef\xbb\xbf")
    lowered = prefix.lower()
    return lowered.startswith(b"<html") or lowered.startswith(b"<!doctype html") or b"<table" in lowered


def _read_html_table_rows(file_path: str | Path) -> list[list[object | None]]:
    raw_bytes = Path(file_path).read_bytes()
    html_text = _decode_html_bytes(raw_bytes)
    soup = BeautifulSoup(html_text, "html.parser")
    table = soup.find("table")

    if table is None:
        raise ValueError("HTML-файл не содержит таблицу")

    rows: list[list[object | None]] = []
    for tr in table.find_all("tr"):
        cells = tr.find_all(["th", "td"])
        if not cells:
            continue

        row: list[object | None] = []
        for cell in cells:
            text = cell.get_text(" ", strip=True)
            row.append(text if text != "" else None)
        rows.append(row)

    return rows


def _normalize_xlrd_cell(cell: xlrd.sheet.Cell, datemode: int):
    if cell.ctype == xlrd.XL_CELL_DATE:
        return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
    if cell.ctype == xlrd.XL_CELL_BOOLEAN:
        return bool(cell.value)
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        numeric_value = float(cell.value)
        if numeric_value.is_integer():
            return int(numeric_value)
        return numeric_value
    if cell.ctype in {xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK}:
        return None
    return cell.value


def _read_excel_rows(file_path: str | Path, sheet_index: int) -> list[list[object | None]]:
    source = Path(file_path)
    suffix = source.suffix.lower()

    if suffix == ".xls":
        raw_bytes = source.read_bytes()
        if _looks_like_html(raw_bytes):
            return _read_html_table_rows(source)

        workbook = xlrd.open_workbook(str(source))
        sheet = workbook.sheet_by_index(sheet_index)
        rows: list[list[object | None]] = []

        for row_index in range(sheet.nrows):
            row: list[object | None] = []
            for column_index in range(sheet.ncols):
                cell = sheet.cell(row_index, column_index)
                row.append(_normalize_xlrd_cell(cell, workbook.datemode))
            rows.append(row)

        return rows

    workbook = openpyxl.load_workbook(str(source), data_only=True)
    worksheet = workbook.worksheets[sheet_index]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def _cast(value, expected_type: str):
    if value is None or str(value).strip() == "":
        raise ValueError("empty")

    t = expected_type.lower()
    if t == "str":
        return str(value)
    if t == "int":
        return int(float(str(value).replace(",", ".")))
    if t == "float":
        return float(str(value).replace(",", "."))
    if t == "date":
        if isinstance(value, datetime):
            return value.date().isoformat()
        raw = str(value).strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%Y-%m-%dT%H:%M:%S"):
            try:
                return datetime.strptime(raw, fmt).date().isoformat()
            except ValueError:
                continue
        raise ValueError("date expected")
    if t == "bool":
        if isinstance(value, bool):
            return value
        low = str(value).strip().lower()
        if low == "да":
            return True
        if low == "нет":
            return False
        raise ValueError("bool expected")
    raise ValueError("unknown type")


def load_xlsx(file_path: str | Path, template: Template) -> tuple[list[dict], list[dict], int]:
    rows = _read_excel_rows(file_path, template.sheet_index)
    if len(rows) <= template.header_row:
        return [], [], 0

    header = [str(c).strip() if c is not None else "" for c in rows[template.header_row]]
    col_index = {name: idx for idx, name in enumerate(header)}

    valid_rows: list[dict] = []
    errors: list[dict] = []
    data_rows = rows[template.header_row + 1 :]

    for row_num, row in enumerate(data_rows, start=template.header_row + 2):
        row_data: dict = {}
        row_valid = True

        for spec in template.columns:
            if spec.name not in col_index:
                if spec.required:
                    errors.append({
                        "Строка": row_num,
                        "Колонка": spec.name,
                        "Ожидаемый тип": spec.expected_type,
                        "Фактическое значение": "<колонка отсутствует>",
                    })
                    row_valid = False
                continue

            raw = row[col_index[spec.name]]

            if spec.required and (raw is None or str(raw).strip() == ""):
                errors.append({
                    "Строка": row_num,
                    "Колонка": spec.name,
                    "Ожидаемый тип": spec.expected_type,
                    "Фактическое значение": "<пусто>",
                })
                row_valid = False
                continue

            if raw is None:
                row_data[spec.name] = None
                continue

            try:
                row_data[spec.name] = _cast(raw, spec.expected_type)
            except Exception:
                errors.append({
                    "Строка": row_num,
                    "Колонка": spec.name,
                    "Ожидаемый тип": spec.expected_type,
                    "Фактическое значение": str(raw),
                })
                row_valid = False

        if row_valid:
            valid_rows.append(row_data)

    return valid_rows, errors, len(data_rows)


def _make_unique_headers(header_row: tuple | list) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}

    for index, cell in enumerate(header_row):
        base = str(cell).strip() if cell is not None and str(cell).strip() else f"Колонка {index + 1}"
        count = seen.get(base, 0)
        seen[base] = count + 1
        headers.append(base if count == 0 else f"{base} ({count + 1})")

    return headers


def load_xlsx_as_is(file_path: str | Path) -> tuple[list[dict], list[str], int]:
    sheet_rows = _read_excel_rows(file_path, 0)

    if not sheet_rows:
        return [], [], 0

    headers = _make_unique_headers(sheet_rows[0])
    rows: list[dict] = []

    for raw_row in sheet_rows[1:]:
        row: dict[str, object] = {}
        for index, header in enumerate(headers):
            row[header] = raw_row[index] if index < len(raw_row) and raw_row[index] is not None else ""
        rows.append(row)

    return rows, headers, len(rows)
