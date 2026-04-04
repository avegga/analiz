import csv
import re
import uuid
from pathlib import Path
from datetime import datetime
from decimal import Decimal, InvalidOperation

import openpyxl
from openpyxl.styles import Font

from src.core.config import EXPORT_DIR


WINDOWS_RESERVED_FILENAMES = {
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
}

MONEY_TYPES = {"money", "денежный"}
MONEY_COLUMN_KEYWORDS = ("сумма", "sum", "amount", "стоимость", "price")


def _normalize_money_value(value: object) -> object:
    if value is None or isinstance(value, (int, float)):
        return value

    raw = str(value).strip()
    if not raw:
        return value

    cleaned = raw.replace("\xa0", " ").replace(" ", "")
    cleaned = re.sub(r"[^0-9,.-]", "", cleaned)
    cleaned = cleaned.strip(".,")
    if not cleaned or cleaned in {"-", ".", ","}:
        return value

    separators = [index for index, char in enumerate(cleaned) if char in {",", "."}]
    normalized = cleaned

    if separators:
        separator_index = separators[-1]
        digits_before = re.sub(r"[.,]", "", cleaned[:separator_index])
        digits_after = re.sub(r"[.,]", "", cleaned[separator_index + 1 :])

        use_decimal_separator = len(separators) > 1 or (0 < len(digits_after) <= 2)
        if not use_decimal_separator and len(digits_after) == 3 and len(digits_before) <= 3:
            use_decimal_separator = True

        if use_decimal_separator:
            normalized = f"{digits_before}.{digits_after}" if digits_after else digits_before
            if cleaned.startswith("-") and not normalized.startswith("-"):
                normalized = f"-{normalized}"
        else:
            normalized = re.sub(r"[.,]", "", cleaned)

    try:
        number = Decimal(normalized)
    except InvalidOperation:
        return value

    if number == number.to_integral_value():
        return int(number)
    return float(number)


def _prepare_rows_for_xlsx(rows: list[dict], column_type_overrides: dict[str, str] | None = None) -> list[dict]:
    if not column_type_overrides:
        column_type_overrides = {}

    money_columns = {
        column_name
        for column_name, column_type in column_type_overrides.items()
        if str(column_type).strip().lower() in MONEY_TYPES
    }

    prepared_rows: list[dict] = []
    for row in rows:
        prepared_row: dict = {}
        for key, value in row.items():
            normalized_key = str(key).strip().lower()
            looks_like_money_column = any(keyword in normalized_key for keyword in MONEY_COLUMN_KEYWORDS)
            should_convert = key in money_columns or looks_like_money_column
            prepared_row[key] = _normalize_money_value(value) if should_convert else value
        prepared_rows.append(prepared_row)
    return prepared_rows


def export_rows(rows: list[dict], fmt: str) -> Path:
    if not rows:
        raise ValueError("Нет данных для экспорта")

    file_id = uuid.uuid4().hex[:10]
    if fmt == "xlsx":
        out_file = EXPORT_DIR / f"export_{file_id}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Данные"
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in rows:
            ws.append([row.get(h, "") for h in headers])
        wb.save(str(out_file))
        return out_file

    if fmt == "csv":
        out_file = EXPORT_DIR / f"export_{file_id}.csv"
        with open(out_file, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys(), delimiter=";")
            writer.writeheader()
            writer.writerows(rows)
        return out_file

    raise ValueError("Неподдерживаемый формат")


def _build_default_export_name() -> str:
    return f"data_{datetime.now().strftime('%Y%m%d_%H%M%S')}"


def _sanitize_filename(filename: str | None) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', "_", (filename or "").strip())
    cleaned = cleaned.rstrip(". ")
    if not cleaned:
        cleaned = _build_default_export_name()

    stem = Path(cleaned).stem.strip() or _build_default_export_name()
    stem = stem.rstrip(". ") or _build_default_export_name()
    if stem.upper() in WINDOWS_RESERVED_FILENAMES:
        stem = f"{stem}_file"
    return f"{stem}.xlsx"


def _build_unique_path(target: Path) -> Path:
    if not target.exists():
        return target

    index = 1
    while True:
        candidate = target.with_name(f"{target.stem}_{index}{target.suffix}")
        if not candidate.exists():
            return candidate
        index += 1


def export_rows_to_xlsx_in_directory(
    rows: list[dict],
    target_dir: str | Path,
    filename: str | None = None,
    column_type_overrides: dict[str, str] | None = None,
) -> Path:
    if not rows:
        raise ValueError("Нет данных для экспорта")

    target = Path(target_dir)
    target.mkdir(parents=True, exist_ok=True)

    out_file = _build_unique_path(target / _sanitize_filename(filename))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Данные"

    prepared_rows = _prepare_rows_for_xlsx(rows, column_type_overrides)

    headers = list(prepared_rows[0].keys())
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in prepared_rows:
        ws.append([row.get(h, "") for h in headers])

    wb.save(str(out_file))
    return out_file
